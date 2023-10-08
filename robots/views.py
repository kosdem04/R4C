from django.shortcuts import render
from django.http import JsonResponse, FileResponse
import json
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from .models import Robot
import random
from openpyxl import Workbook

# Представление для первого задания
# @csrf_exempt — декоратор в Django, который используется для отключения проверки CSRF
@csrf_exempt
def robots(request):
    # Проверяем тип запроса
    if request.method == 'POST':
        try:
            # Загружаем информацию в формате JSON из запроса
            data = json.loads(request.body)
            # Передаём полученную информацию в модель для последующей проверки
            robots = Robot(**data)
            # Проверяем входные данные, в случае ошибки происходит исключение ValidationError
            robots.full_clean()
            # Делаем проверку, есть ли робот с входным серийным номером в базе данных
            for rob in Robot.objects.all():
                if robots.serial == rob.serial:
                    print('Робот с таким серийным номером уже есть в базе данных')
                    return JsonResponse({'message':'Робот с таким серийным номером уже есть в базе данных'})
            # Сохраняем информацию в базе данных
            robots.save()
            return JsonResponse({'message': 'Robot created successfully'}, status=200)
        # Обработка исключения ValidationError
        except ValidationError as error:
            return JsonResponse({'error': str(error)}, status=400)
    # Обработка GET-запроса
    else:
        return JsonResponse({'message':'Нет активных функций для GET-запроса'})

# Представления для второго задания
# Вывод ссылки для скачивания Exel-файла
def exel_file(request):
    return render(request, 'upload_exel.html')

# Скачивание Exel-файла
def upload_exel(request):
    robots = Robot.objects.all()
    # Определяем информацию, которая будет записана в Exel-файл путём запроса к базе данных и рандомайзера
    data = []
    for robot in robots:
        data += [{'model': robot.model, 'version': robot.version, 'количество за неделю': random.randint(10, 100)}]
    # Создание новой Exel-файла
    exel = Workbook()
    # Создаём для каждой модели свою страницу при помощи множества
    for model in set([row['model'] for row in data]):
        # Создаём заголовок для листа
        sheet = exel.create_sheet(title=f'Модель {model}', index=0)
        # Прописываем заголовки
        sheet['A1'] = 'Модель'
        sheet['B1'] = 'Версия'
        sheet['C1'] = 'Количество за неделю'
        # Заполняем лист информацией
        for i, row in enumerate([row for row in data if row['model'] == model], start=2):
            sheet.cell(row=i, column=1, value=row['model'])
            sheet.cell(row=i, column=2, value=row['version'])
            sheet.cell(row=i, column=3, value=row['количество за неделю'])
    # Прописываем путь и прописываем автоматическое имя Exel-файла
    excel_path = f'exel/model_value_for_week{random.randint(10, 100)}.xlsx'
    # Сохраняем Exel-файл
    exel.save(excel_path)
    # Открываем Exel-файл
    excel_file = open(excel_path, 'rb')
    # Создаём HTTP-ответ с Exel-файлом в качестве содержимого
    response = FileResponse(excel_file, as_attachment=True, filename=f'model_value_for_week{random.randint(10, 100)}.xlsx')
    return response
